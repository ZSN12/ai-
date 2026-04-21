# utils.py
# 文件读取、行数估算等工具函数

import pandas as pd
import re
from typing import Optional, List, Any

def estimate_csv_rows(file_obj) -> int:
    """快速估算CSV行数（不含表头）"""
    file_obj.seek(0)
    total_bytes = file_obj.getbuffer().nbytes if hasattr(file_obj, 'getbuffer') else len(file_obj.getvalue())
    if total_bytes == 0:
        return 0
    sample_size = min(100000, total_bytes)
    sample_bytes = file_obj.read(sample_size)
    file_obj.seek(0)
    try:
        sample_text = sample_bytes.decode('utf-8')
    except UnicodeDecodeError:
        sample_text = sample_bytes.decode('gbk', errors='ignore')
    sample_lines = sample_text.count('\n')
    if sample_lines == 0:
        return 0
    avg_bytes_per_line = sample_size / sample_lines
    estimated_total_lines = total_bytes / avg_bytes_per_line
    return max(0, int(estimated_total_lines) - 1)  # 减去表头

def estimate_excel_rows(file_obj) -> int:
    """使用 openpyxl 只读模式估算 Excel 行数"""
    try:
        from openpyxl import load_workbook
        file_obj.seek(0)
        wb = load_workbook(filename=file_obj, read_only=True)
        sheet = wb.active
        # 使用 calculate_dimension 获取实际使用的范围，比 max_row 更准确
        dimension = sheet.calculate_dimension()
        if dimension and ':' in dimension:
            # 解析维度字符串，如 "A1:Z1000"
            end_row = int(dimension.split(':')[1].lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))
            wb.close()
            return max(0, end_row - 1) if end_row else 0
        # 如果 calculate_dimension 无效，尝试从末尾遍历找到最后一行
        rows = sheet.max_row
        wb.close()
        return max(0, rows - 1) if rows else 0
    except Exception as e:
        print(f"估算Excel行数失败: {str(e)}")
        return -1

def estimate_rows(file_obj) -> int:
    """估算上传文件的总数据行数（不含表头）"""
    filename = file_obj.name.lower()
    file_obj.seek(0)
    if filename.endswith('.csv'):
        return estimate_csv_rows(file_obj)
    elif filename.endswith(('.xlsx', '.xls')):
        return estimate_excel_rows(file_obj)
    else:
        return 0

def read_any_file(file_obj, chunk_size: Optional[int] = None) -> pd.DataFrame:
    """读取 CSV 或 Excel 文件，强制 Type 列为 int32"""
    if file_obj is None:
        return pd.DataFrame()
    file_name = file_obj.name.lower()
    file_obj.seek(0)
    try:
        if file_name.endswith('.csv'):
            if chunk_size:
                chunks = []
                try:
                    for chunk in pd.read_csv(file_obj, chunksize=chunk_size, encoding='utf-8', on_bad_lines='skip'):
                        if 'Type' in chunk.columns:
                            chunk['Type'] = pd.to_numeric(chunk['Type'], errors='coerce').fillna(0).astype('int32')
                        chunks.append(chunk)
                except Exception as e:
                    print(f"读取CSV文件失败: {str(e)}")
                    return pd.DataFrame()
                if not chunks:
                    return pd.DataFrame()
                df = pd.concat(chunks, ignore_index=True)
            else:
                try:
                    df = pd.read_csv(file_obj, encoding='utf-8')
                except Exception as e:
                    print(f"读取CSV文件失败: {str(e)}")
                    return pd.DataFrame()
        else:
            try:
                if chunk_size:
                    # 对于Excel文件，使用openpyxl引擎进行分块读取
                    from openpyxl import load_workbook
                    file_obj.seek(0)
                    wb = load_workbook(filename=file_obj, read_only=True)
                    sheet = wb.active
                    chunks = []
                    rows = []
                    header = None
                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if i == 0:
                            header = row
                        else:
                            rows.append(row)
                            if len(rows) >= chunk_size:
                                chunk_df = pd.DataFrame(rows, columns=header)
                                if 'Type' in chunk_df.columns:
                                    chunk_df['Type'] = pd.to_numeric(chunk_df['Type'], errors='coerce').fillna(0).astype('int32')
                                chunks.append(chunk_df)
                                rows = []
                    if rows:
                        chunk_df = pd.DataFrame(rows, columns=header)
                        if 'Type' in chunk_df.columns:
                            chunk_df['Type'] = pd.to_numeric(chunk_df['Type'], errors='coerce').fillna(0).astype('int32')
                        chunks.append(chunk_df)
                    if chunks:
                        df = pd.concat(chunks, ignore_index=True)
                    else:
                        df = pd.DataFrame()
                else:
                    df = pd.read_excel(file_obj)
            except Exception as e:
                print(f"读取Excel文件失败: {str(e)}")
                return pd.DataFrame()

        if 'Type' in df.columns:
            df['Type'] = pd.to_numeric(df['Type'], errors='coerce').fillna(0).astype('int32')
        return df
    except Exception as e:
        print(f"读取文件失败: {str(e)}")
        return pd.DataFrame()

def sanitize_filename(filename: str) -> str:
    """清理文件名，移除不安全字符"""
    # 移除路径遍历字符
    filename = filename.replace('..', '').replace('/', '').replace('\\', '')
    # 只保留安全字符
    filename = re.sub(r'[^\w\-.]', '_', filename)
    # 限制文件名长度
    return filename[:255]

def validate_file_extension(filename: str, allowed_extensions: List[str]) -> bool:
    """验证文件扩展名是否在允许的列表中"""
    if not filename:
        return False
    ext = filename.lower().split('.')[-1] if '.' in filename else ''
    return ext in [ext.lower() for ext in allowed_extensions]

def validate_file_size(file_size: int, max_size_mb: int) -> bool:
    """验证文件大小是否在允许的范围内"""
    max_size_bytes = max_size_mb * 1024 * 1024
    return file_size <= max_size_bytes

def clean_text(text: str) -> str:
    """清理文本，移除潜在的危险字符"""
    if not text:
        return ""
    # 移除控制字符（除了换行和制表符）
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    # 规范化空白字符
    text = ' '.join(text.split())
    return text.strip()

def validate_dataframe_columns(df: pd.DataFrame, required_columns: List[str]) -> tuple[bool, List[str]]:
    """验证数据框是否包含必需的列"""
    if df is None or df.empty:
        return False, ["数据框为空"]
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return False, [f"缺少必需列: {', '.join(missing_columns)}"]
    
    return True, []

def safe_convert_to_int(value: Any, default: int = 0) -> int:
    """安全地将值转换为整数"""
    try:
        if pd.isna(value):
            return default
        return int(float(value))
    except (ValueError, TypeError):
        return default

def safe_convert_to_str(value: Any, default: str = "") -> str:
    """安全地将值转换为字符串"""
    try:
        if pd.isna(value):
            return default
        return str(value).strip()
    except Exception:
        return default

def format_file_size(size_bytes: int) -> str:
    """格式化文件大小为人类可读格式"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} TB"

def is_valid_email(email: str) -> bool:
    """验证邮箱地址格式"""
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def is_valid_phone(phone: str) -> bool:
    """验证手机号格式（简单的格式检查）"""
    if not phone:
        return False
    # 移除所有非数字字符
    phone_digits = re.sub(r'\D', '', phone)
    # 检查是否为11位数字（中国手机号）
    return len(phone_digits) == 11 and phone_digits.isdigit()